# -*- coding: utf-8 -*-
"""
Created on Thu Dec  7 22:02:26 2017

@author: Administrator
"""

import pandas as pd
from geopy.distance import vincenty
from openpyxl import load_workbook


railway = pd.read_excel("./Railway_stations_Melbourne.xlsx", sheetname='Sheet1')


wb = load_workbook(filename = 'RailwayStationCal.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
distance = []

for row in range(2, 98379):
    prop_lat = ws['H' + str(row)].value
    prop_long = ws['I' + str(row)].value
    prop_loc = (prop_lat, prop_long)
    distance = []
    
    for i in railway.index:
        rail_lat = railway.Latitude[i]
        rail_long = railway.Longitude[i]
        rail_loc = (rail_lat, rail_long)
        dist = vincenty(prop_loc, rail_loc).meters
        distance.append(dist)
    
    dist = sorted(distance)[0]
    ws['J' + str(row)].value = dist
    distance[:] = []
    
    print(row, dist)

      

wb.save('RailwayStation.xlsx') 

