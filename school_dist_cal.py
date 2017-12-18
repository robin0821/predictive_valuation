

import pandas as pd
from geopy.distance import vincenty
from openpyxl import load_workbook


primary = pd.read_excel("./Primary.xlsx", sheetname='Sheet1')



secondary = pd.read_excel("./secondary.xlsx", sheetname='Sheet1')



wb = load_workbook(filename = 'real_estate_mel_cln.xlsx')
ws = wb.get_sheet_by_name('Sheet1')

for row in range(1, 18385):
    prop_lat = ws['H' + str(row)].value
    prop_long = ws['I' + str(row)].value
    prop_loc = (prop_lat, prop_long)
    distance = []
    
    for i in primary.index:
        sch_lat = primary.Latitude[i]
        sch_long = primary.Longitude[i]
        sch_loc = (sch_lat, sch_long)
        dist = vincenty(prop_loc, sch_loc).meters
        distance.append(dist)
    
    primary['distance'] = distance
    a = primary.nsmallest(3, 'distance').index
    ws['Q' + str(row)].value = primary.SchoolName[a[0]]
    ws['R' + str(row)].value = primary.Sector[a[0]]
    ws['S' + str(row)].value = primary.distance[a[0]]
    ws['T' + str(row)].value = primary.Score[a[0]]
    ws['U' + str(row)].value = primary.SchoolName[a[1]]
    ws['V' + str(row)].value = primary.Sector[a[1]]
    ws['W' + str(row)].value = primary.distance[a[1]]
    ws['X' + str(row)].value = primary.Score[a[1]]
    ws['Y' + str(row)].value = primary.SchoolName[a[2]]
    ws['Z' + str(row)].value = primary.Sector[a[2]]
    ws['AA' + str(row)].value = primary.distance[a[2]]
    ws['AB' + str(row)].value = primary.Score[a[2]]
    print(row, primary.SchoolName[a[0]], primary.distance[a[0]])
    
#    if row % 1000 == 0:
#        wb.save('real_estate_mel_cln.xlsx')
    
    distance[:] = []
    for i in secondary.index:
        sch_lat = secondary.Latitude[i]
        sch_long = secondary.Longitude[i]
        sch_loc = (sch_lat, sch_long)
        dist = vincenty(prop_loc, sch_loc).meters
        distance.append(dist)
    
    secondary['distance'] = distance
    a = secondary.nsmallest(3, 'distance').index
    ws['AC' + str(row)].value = secondary.SchoolName[a[0]]
    ws['AD' + str(row)].value = secondary.Sector[a[0]]
    ws['AE' + str(row)].value = secondary.distance[a[0]]
    ws['AF' + str(row)].value = secondary.Score[a[0]]
    ws['AG' + str(row)].value = secondary.SchoolName[a[1]]
    ws['AH' + str(row)].value = secondary.Sector[a[1]]
    ws['AI' + str(row)].value = secondary.distance[a[1]]
    ws['AJ' + str(row)].value = secondary.Score[a[1]]
    ws['AK' + str(row)].value = secondary.SchoolName[a[2]]
    ws['AL' + str(row)].value = secondary.Sector[a[2]]
    ws['AM' + str(row)].value = secondary.distance[a[2]]
    ws['AN' + str(row)].value = secondary.Score[a[2]]
    print(row, secondary.SchoolName[a[0]], secondary.distance[a[0]])
    
#    if row % 1000 == 0:
#        wb.save('real_estate_mel_cln.xlsx')    

wb.save('real_estate_mel_cln.xlsx') 




