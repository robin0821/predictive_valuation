# -*- coding: utf-8 -*-
"""
Created on Mon Dec  4 09:59:45 2017

@author: Administrator
"""

from openpyxl import load_workbook
from geopy.distance import vincenty

wb = load_workbook(filename = 'real_estate_mel_fin.xlsx')
ws = wb.get_sheet_by_name('Sheet1')

flinder_lat = -37.818078
flinder_long = 144.96681
flinders_station = (flinder_lat, flinder_long)

for row in range(98001, 98379):
    latitude = ws['H' + str(row)].value
    longitude = ws['I' + str(row)].value
    house_location = (latitude, longitude)
    distance = vincenty(flinders_station, house_location).kilometers
    
    ws['P' + str(row)].value = distance
       
    print(row, distance)
wb.save('real_estate_mel_fin.xlsx')
print("done")    
#    if row % 1000 == 0:
#        wb.save('real_estate_mel_cln.xlsx')
